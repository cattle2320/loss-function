import numpy as np
from tools.list_mean import list_mean
import matplotlib.pyplot as plt
import pandas as pd
import cfg
from openpyxl import Workbook
# fcn
ce_1=[0.9631122 ,0.93132433,0.03257432,0.98530444,0.78355266,0.78282429,0.32904416,0.28523679,0.83718107,0.19937046,0.11121822,0.60721964]
wce1_2= [0.90331695,0.68865285,0.51010367,0.89781251,0.82396862,0.68299483,0.80816805,0.37870086,0.86003906,0.73770385,0.5080597 ,0.53324385]
topk_2=[0.96307039,0.9318538 ,0.03713474,0.98584963,0.78844684,0.78409156,0.3332936 ,0.2960551 ,0.8338218 ,0.21505317,0.13250346,0.61721501]
focal_3=[0.96339122,0.93124776,0.06233925,0.98488746,0.79345198,0.78354804,0.37669333,0.30202131,0.840442  ,0.24367655,0.13060171,0.62944208]
DPCE_4=[0.96363489,0.93209466,0.03186617,0.98572794,0.77136266,0.78393764,0.28685341,0.27728461,0.82701729,0.19122394,0.09913549,0.60491281]
dice_5=[0.94655924,0.90789301,0.1376738 ,0.9778201 ,0.812423  ,0.7492831,0.41114682,0.36060609,0.83911579,0.49994983,0.35492301,0.54556676]
lovasz_6=[0.93451675,0.87407534,0.4458763 ,0.97256535,0.83384465,0.7502171,0.75924603,0.        ,0.88304434,0.71091777,0.        ,0.64871354]
tversky_7=[0.93721811,0.89745082,0.1086354 ,0.96901848,0.85682168,0.76700752,0.50333579,0.38137256,0.87743164,0.44242994,0.31004825,0.6241242 ]
GDice_8=[0.93923134,0.90345168,0.14831916,0.97758946,0.8071966 ,0.75038775,0.42969133,0.34968466,0.84432166,0.48239087,0.31664333,0.57403965]
asym_9=[0.94830067,0.89055059,0.11702674,0.96897886,0.84893249,0.75699036,0.44674418,0.38687575,0.87043608,0.56427619,0.40919786,0.58708098]

PGDice_10=[]
boundary_11=[]
hausdorff_12 = []
sensitivity_specificity_loss13 = [0.9427661 ,0.88066108,0.59300955,0.96880376,0.87692063,0.76833879,0.86761879,0.43679296,0.91878996,0.78774783,0.53381493,0.68728877]
focal_tversky_loss14 = [0.956307  ,0.90645344,0.        ,0.97347835,0.85477347,0.76398649,0.46439755,0.33762223,0.88008982,0.53499008,0.34080639,0.54797416]
overall_Mismatch15 = [0.94976441,0.87716768,0.5302482 ,0.97310532,0.89100895,0.74640632,0.82489302,0.43835963,0.92257852,0.75947939,0.50314218,0.66522453]


# unet
# ce_1=[0.97751341,0.95563706,0.14773289,0.98840585,0.88850735,0.82257223,
# 0.61663774,0.35187356,0.91852378,0.39978079,0.05418813,0.92935796]
#
# wce1_2= [0.90331695,0.68865285,0.51010367,0.89781251,0.82396862,0.68299483,0.80816805,0.37870086,0.86003906,0.73770385,0.5080597 ,0.53324385]
#
# topk_2=[0.98275866,0.8129307 ,0.17485934,0.98827851,0.89819533,0.81608774,
# 0.59834693,0.31228642,0.91301931,0.41866981,0.00799745,0.94129532]
#
# focal_3=[0.9788723 ,0.96002481,0.11909255,0.9894183 ,0.88436273,0.81587245,
# 0.49797603,0.25967331,0.91035158,0.27315109,0.18599868,0.92963593]
#
# DPCE_4=[0.96363489,0.93209466,0.03186617,0.98572794,0.77136266,0.78393764,0.28685341,0.27728461,0.82701729,0.19122394,0.09913549,0.60491281]
# dice_5=[0.94655924,0.90789301,0.1376738 ,0.9778201 ,0.812423  ,0.7492831,0.41114682,0.36060609,0.83911579,0.49994983,0.35492301,0.54556676]
# lovasz_6=[0.93451675,0.87407534,0.4458763 ,0.97256535,0.83384465,0.7502171,0.75924603,0.        ,0.88304434,0.71091777,0.        ,0.64871354]
# tversky_7=[0.93721811,0.89745082,0.1086354 ,0.96901848,0.85682168,0.76700752,0.50333579,0.38137256,0.87743164,0.44242994,0.31004825,0.6241242 ]
# GDice_8=[0.93923134,0.90345168,0.14831916,0.97758946,0.8071966 ,0.75038775,0.42969133,0.34968466,0.84432166,0.48239087,0.31664333,0.57403965]
# asym_9=[0.94830067,0.89055059,0.11702674,0.96897886,0.84893249,0.75699036,0.44674418,0.38687575,0.87043608,0.56427619,0.40919786,0.58708098]
#
# PGDice_10=[]
# boundary_11=[]
# hausdorff_12 = []
#
# sensitivity_specificity_loss13 =[0.94344364,0.87195815,0.8769885 ,0.96127876,0.9122376 ,0.78220136,
#                                  0.89185458,0.4478688 ,0.94263474,0.82250688,0.50800322,0.79892765]
#
# focal_tversky_loss14 = [0.956307  ,0.90645344,0.        ,0.97347835,0.85477347,0.76398649,0.46439755,0.33762223,0.88008982,0.53499008,0.34080639,0.54797416]
#
# overall_Mismatch15 = [0.98561547,0.82593645,0.71958712,0.98480816,0.92126817,0.69807619
# ,0.86120153,0.43996233,0.94597952,0.80663239,0.47201408,0.85459072]



def nanmean(list):
    for i, data in enumerate(list):
        print('{}------------{}'.format(list_name[i], list_mean(data)))

def Transpose(list):
    # 数组转置
    data_all = np.zeros(shape=(len(list), len(list[0])))
    for i, data in enumerate(list):
        if len(data) !=0:
            data_all[i] = data
    num = np.array(data_all)
    # zhuanzhi = num.transpose()
    return num

def plot(x,y,linestyle, label):
    plt.plot(x, y, linestyle, label=label)
    plt.legend()        # 显示上面的label

def ply(data, name, style):
    # x_axis_data = [i for i in range(len(data[0]))]  # x
    x_axis_data = name

    plot(x=x_axis_data, y=data[0], linestyle=style, label='CE Loss')
    plot(x=x_axis_data, y=data[1], linestyle=style, label='Topk Loss')
    plot(x=x_axis_data, y=data[2], linestyle=style, label='Focal Loss')
    plot(x=x_axis_data, y=data[3], linestyle=style, label='DPCE Loss')
    plot(x=x_axis_data, y=data[4], linestyle=style, label='Dice Loss')
    plot(x=x_axis_data, y=data[5], linestyle=style, label='Lovasz Loss')
    plot(x=x_axis_data, y=data[6], linestyle=style, label='Tversky Loss')
    plot(x=x_axis_data, y=data[7], linestyle=style, label='GDice Loss')
    plot(x=x_axis_data, y=data[8], linestyle=style, label='Asym Loss')
    # plot(x=x_axis_data, y=data[9], linestyle=style, label='PGDice loss')
    # plot(x=x_axis_data, y=data[10], linestyle=style, label='Boundary loss')
    # plot(x=x_axis_data, y=data[11], linestyle=style, label='hausdorff_12')
    plot(x=x_axis_data, y=data[12], linestyle=style, label='Sensitivity Specificity Loss')
    plot(x=x_axis_data, y=data[13], linestyle=style, label='Focal Tversky Loss')
    plot(x=x_axis_data, y=data[14], linestyle=style, label='Generalized Region Loss')
    # plot(x=x_axis_data, y=data[15], linestyle=style, label='Wce Loss')

    plt.xlabel('Category')               # x_label
    plt.ylabel('Accuracy')                    # y_label

    plt.xticks(ticks=x_axis_data)
    # plt.yticks(ticks=[i for i in range(0, max)])      # 不显示所有y刻度标签

    plt.grid(True)                  # 显示网格线

    # plt.ylim(-1,1)                #仅设置y轴坐标范围
    plt.show()

def cfg_name(file_path):
    pd_label_color = pd.read_csv(file_path, sep=',')
    colormap = []
    for i in range(len(pd_label_color.index)):
        tmp = pd_label_color.iloc[i]
        name = tmp['name']
        colormap.append(name)
    return colormap

if __name__ == '__main__':
    list_name = ['ce_1', "topk_2", "focal_3", "DPCE_4", "dice_5", "lovasz_6", "tversky_7", "GDice_8", "asym_9", "PGDice_10", "boundary_11",
        "hausdorff_12", "sensitivity_specificity_loss13", "focal_tversky_loss14", "overall_Mismatch15", 'wce1_2']

    list1_name = ['CE', "Topk", "Focal", "DPCE", "Dice", "Lovasz", "Tversky", "GDice", "Asym",
                 "SS", "Focal Tversky",
                 'Wce',"Generalized Region"]
    list1 = [ce_1, topk_2, focal_3, DPCE_4, dice_5, lovasz_6, tversky_7, GDice_8, asym_9, sensitivity_specificity_loss13, focal_tversky_loss14, wce1_2, overall_Mismatch15]

    list = [ce_1, topk_2, focal_3, DPCE_4, dice_5, lovasz_6, tversky_7, GDice_8, asym_9, PGDice_10, boundary_11,
        hausdorff_12, sensitivity_specificity_loss13, focal_tversky_loss14, overall_Mismatch15, wce1_2]

    Transpose_data = Transpose(list1)
    print(Transpose_data)
    max_idex = []
    for i in Transpose_data:
        p = np.argmax(i)
        max_idex.append(p+1)
    print(max_idex)

    name = cfg_name(r'D:\Desktop\FCN-master\Datasets\CamVid\class_dict.csv')


    # 将数据写入excel
    from openpyxl import load_workbook
    # 加载excel，注意路径要与脚本一致
    wb = load_workbook(r'D:\Desktop\123456.xlsx')
    # 激活excel表
    sheet = wb.active

    for n in range(len(name)):
        sheet.cell(row=1, column=n+2).value = name[n]

    for n in range(len(list1_name)):
        sheet.cell(row=n+2, column=1).value = list1_name[n]

    for i in range(13):
        for j in range(12):
            data = round(Transpose_data[i][j], 4)
            sheet.cell(row=i+2, column=j+2).value = data

    wb.save(r'D:\Desktop\123456.xlsx')

    print('数据写入成功！')


    # for i in range(len(name)):
    #     list[i] = Transpose_data[i]

    # ply([data for data in list], name, style='-.')



